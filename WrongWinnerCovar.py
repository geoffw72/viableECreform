# This program reads in a list of states, their EVs, and the mean and std dev of their departure from natl vote
#  These departures are correlated via a covariance matrix.  The user specifies the noise level.
# The program generates a large number of state departure sets.  For each set, the program
# calculates (for each h value) the national vote that would generate a toss-up = Vustu.
# It then simulates elections around that toss-up, to determine the fraction of elections for each h
# that would generate a "missed election" - incorrect national outcome for the supplied noise
#    The outputs (by h value) are the mean toss-up vote, and the fraction of elections won by the wrong
#    candidate (both at a given distance from the toss-up, and integrated over the competitive range)
#  The integrations are done only for a uniform distribution
# Note: all my code expresses Vi in terms of Dem's share, while paper uses GOP as candidate 1


from openpyxl import load_workbook  # **PREPPING TO READ IN THE STATE DATA
from openpyxl.utils import get_column_letter, column_index_from_string
wb = load_workbook(filename = "state_data.xlsx")
ws = wb.active
import numpy as np
from scipy.stats import norm
import xlwt
from xlwt import Workbook
wb = Workbook()
nationalAvg = 0.50
minNatlAvg = input("What is the min. national average fraction vote for Dems, e.g. 0.478? ")
maxNatlAvg = input("What is the max national average fraction vote for Dems, e.g. 0.538? ")
numberVus = input("How many Vus to simulate? (e.g. 61) ")
kappa = 1.001 #ignore this; only used for gamma distribution code
theta = 0.099 #ignore this; only used for gamma distribution code
nStates=51  # Excel data in rows B, C, D, E starting in row2
print("number of states = "+str(nStates)+". Time to read in state data.")
nonnoise = [0.]*nStates #initial the election-specific state dev'n from nat'l avg (local to loop)
corrldNonnoise = [0.]*nStates #initial the correlated state dev'n from nat'l avg (local to loop)
error = [0.]*nStates  #initialize the election+state-specific error tuple (local to loop)
maxEV = 0   # this will be total US available electoral college votes (global variable)
stateName = ["xx"]*nStates #initializing state's name, no of EVs, mean di, std-dev
stateEV = [0.]*nStates
avgDelta = [0.]*nStates
stateSD = [0.]*nStates

for j in range(nStates) :   #READ IN THE NATIONAL LANDSCAPE (state name, state EV, avgDelta, stateSD)
    currentRow = str(j+2)  #first row of this file is a header row
    currentStateNameCell = "B"+currentRow
    currentStateName = ws[currentStateNameCell]
    currentStateEVCell = "C" + currentRow
    currentStateEV = ws[currentStateEVCell]
    currentStateADCell = "D" + currentRow
    currentStateAD = ws[currentStateADCell]
    currentStateSDCell = "E" + currentRow
    currentStateSD = ws[currentStateSDCell]
    maxEV += currentStateEV.value  # add this state's EV max to the US max
   # assign this state's properties
    stateName[j] =currentStateName.value
    stateEV[j] = currentStateEV.value
    avgDelta[j] = float(currentStateAD.value)  #these are the <di> mean departures
    stateSD[j] = currentStateSD.value          #these are the si (std dev in di)

print("The total number of EV's available in the US is "+str(maxEV))
Sims = input("How many simulations to run per Vus?  reco = 40000 ")
noSims = int(Sims)
print("Each state's Dem EV's will be a tilted step between 0.5-d and 0.5+d")
# print("I will compute election results for five d values, including d=0=EC")
nBetas = int(8)
nEV_mean = [[0]*nBetas for nVus in range(numberVus) ]   #the Dems get DC for an average election for all h's
sumTotalEV = [0.]*nBetas  #new variable for 2024.  Running average EV award for all h
sumVarTotalEV = [0.]*nBetas  #new variable for 2024. Running average variance for all h
totalEV = [0.]*nBetas #initialize electoral votes for each h
totalEV_noisy = [0.]*nBetas #initialize noisy national EV's as well
stdevEV = [0.]*nBetas   #initialize std dev in EV's for each h

Vustu = [[0.5]*nBetas for i in range(noSims) ]  #this tracks the toss-up point for each sim
nDemWins = [[0]*nBetas for nVus in range(numberVus) ] # the Dems haven't won any elections yet for h=9999
nNoisyDemWins = [[0]*nBetas for nVus in range(numberVus) ]
nMissedDemWins = [[0]*nBetas for nVus in range(numberVus) ] # no wins either when noise has boosted into win
nMissedDemLosses =[[0]*nBetas for nVus in range(numberVus) ]  # no wins when noise has sagged into loss
nTotDW = [0]*nBetas   #this row and below three are for totalling wins and missed wins across all sims
nTotNDW = [0]*nBetas
nTotMDW = [0]*nBetas
nTotMDL = [0]*nBetas
nDeltas = nBetas
beta=[9999,50,20,14,10,6,3,1]  #beta = h.  h=9999 is flag for current EC
delta = [0]*nDeltas   #delta is convenient shorthand for EV award calcns

for k in range(nBetas) :
    delta[k]=0.50/beta[k]    #shorthand expression for 0.50/h; useful for below calcns

deltaVus = (maxNatlAvg - minNatlAvg)/(numberVus - 1.0) #step size in looped national vote
noise = input("What is the stdev in noise per state? (0.001, .0017, 0.003, 0.005, 0.01) ")
outName = "{:.4f}".format(noise)  #for naming the outputfile
# fudgeFactor = input("input the fudgeFactor for reducing correlated noise (e.g. 2.0) ")  #obsolete

coVar = [[0] * nStates for j in range(nStates)] # this initializes the Cholesky for covariance
print("I will now read in the Cholesky decomposition of the covariance matrix")
for j in range(nStates) :
    currentRow = str(j+2)
    for jj in range(nStates) :
        CL = get_column_letter(jj+7)
        dataLoc = ws[CL+currentRow]
        dataValue = float(dataLoc.value)
        print(currentRow, CL,dataValue)
        coVar[j][jj]=dataValue  #assigning the covariance Cholesky matrix
print("I will now run "+str(noSims)+" simulated elections for several V_EWV levels")

EVtot = [[0] * nBetas for i in range(noSims) ]  #NEW: initialize a table of results
#NEW - let's calculate some mean results for this avg {d} set
n_samples = 200
meanVustu = [0.5]*nBetas #this will be the toss-up Vus for this h and the mean {di} set
stdevVustu = [0.]*nBetas  #std dev in toss-up Vus(h)
localVustu = [[0.]*nBetas for i in range(n_samples)]  #list of Vustu's found in below search routine
for nVus in range(numberVus) :
    nationalAvg = minNatlAvg+deltaVus*nVus  #set nationalAvg to current Vus

    # QUALITY CHECK: Compute the Electoral College result for each h for each state's mean Vi  
    for k in range(nBetas) :
        nEV_mean[nVus][k]=0  #reset the Dem's EV award to DC only b4 looping thru states
        for j in range(nStates) :
            p = nationalAvg + avgDelta[j] #local variable for this state's Vi
            if (beta[k] >= 900) :      #we're at h = infinity; delta=0
                award = 0   #default = GOP win
                if p > 0.50 :
                    award = stateEV[j]   #holy cow! Dems win this state
            else : #this is not the EC case, finite beta = h            
                award = stateEV[j]*min( 1,max(0,0.5*(1+beta[k]*(p/0.5 - 1))) )
            nEV_mean[nVus][k] += award       #add these earned votes for this state to this beta's total
            # print("B,state no,p,award "+str(beta[k])+","+stateName[j]+","+str(p)+","+str(award))
            strB = str(beta[k])
            demEVs = str(round(nEV_mean[nVus][k],4))
        print("we are done computing EC results at natl avg Vus = "+str(nationalAvg))
        print("at 0.5/delta = "+strB+", "+demEVs+" = Dem EV's in absence of state variation, noise")
        
#BELOW: CODE BLOCK TO ITERATE to find avgVustu for this f(di) set.  Do for 200 samples
#  this code block eliminated for covariance case -- not shared as an output
    # END OF CODE BLOCK TO FIND 200 samples of Vustu FOR EACH VALUE OF h

    # NOW COMPUTE MEAN AND STD DEV IN VUSTU ACROSS THESE 200 SAMPLES
    localSumVar = 0.
    for i in range(n_samples) :  #mini loop to calc std dev in Vustu for this h value
        localSumVar += (localVustu[i][k]-meanVustu[k])**2
    stdevVustu[k] = ( localSumVar/(n_samples-1.0) )**0.5
    printPB = "{:.4f}".format(meanVustu[k]-0.5)  #these three lines are for debugging ...
    printPBS = "  " + "{:.5f}".format(stdevVustu[k])
    print("h, mean and stdev in partisan bias are "+str(beta[k])+", "+printPB+printPBS)
    

#Before we start the sims for missed-election calcn, let's zero out some stat counters
for k in range(0,nBetas) :  #zero out the counters for this value of Vus
    sumTotalEV[k]=0.
    sumVarTotalEV[k]=0.
    totalEV[k] = 0   #fill in zeros for total EVs.
    totalEV_noisy[k]=0

# NOW RUN THE MAIN SIMULATIONS; all preliminaries and mean values have been calculated

for nVus in range(numberVus) : #each Vus interval ("bin") gets its own worksheet
    Vus = minNatlAvg+nVus*deltaVus  #setting the mean natl avg for this group of sims
    sumTotalEV=[0.]*nBetas #reset this counter for this Vus
    for i in range(noSims) :  #run a number of simulations near this Vus value
        # For each sim, assign each state's Vi informed by the f(Vi) distribution
        sumOffset = 0.  #start counter for random drift from sum[EVtot(j)*di(j)] must = 0
        for j in range(nStates) :  #loop to tentatively assign the delVi's
            nonnoise[j] = np.random.normal(0.0,stateSD[j])  #here, we sample the (di,sigma-i) space
            error[j] = np.random.normal(0.0,noise)
            sumOffset += stateEV[j]*( avgDelta[j]+nonnoise[j] )  #will be used in below normalization
        #Now, tweak the nonnoise to force normalization.  We do not normalize the error set
        for j in range(nStates) :
            nonnoise[j] = nonnoise[j]-sumOffset/(maxEV)  #correcting for bias in overall state dev'ns
        
        # NEW for FULL COVARIANCE CODE - implement covariance among state wobbles from <di>
        corrldNonnoise = np.dot(coVar,nonnoise)  #let's try having nonnoise be unit s.d.
        sumCorrldNN = 0.  #counter for offset in sum(di*EVi) from our correlation matrix
        for j in range(nStates) :
            sumCorrldNN += stateEV[j]*corrldNonnoise[j]    #add to running total off-center total of di's
        for j in range(nStates) :
            nonnoise[j] = corrldNonnoise[j] - sumCorrldNN/maxEV #Rezeroing again to center the mean
            # we used to apply a fudge factor to the variance here, but that has been eliminated
            # corrldNonnoise[j] = corrldNonnoise[j] / fudgeFactor  #obsolete

        for k in range(nBetas) :
            adjustVus = deltaVus*(float(i)/float(noSims)-0.5)  #so we can sample uniformly in this interval
            nationalAvg = Vus + adjustVus  #for each k, assign the national avg
            totalEV[k] = 0       # Reset how many EVs the Dems have before counting Alabama
            totalEV_noisy[k] = 0  #default is GOP won them all
            for j in range(nStates) : #now loop over states to calc awards for this h value
                Vi = nationalAvg + avgDelta[j] + nonnoise[j]  #calc this state's Vi - local loop variable
                noisyVi = Vi + error[j]
                award = 0.  #default = GOP swept the state
                if(beta[k] < 900) :   #we're not doing the h=infinity case. for noisy and non, are we inside step?
                    if Vi > 0.50 - delta[k] :  #Dems got at least some nonnoisy EV's
                        award = stateEV[j] #all to Dems unless split
                        if Vi < 0.50 + delta[k] :   #split award for non-noisy election
                            award = 0.5 * stateEV[j] * (1 + beta[k]*(Vi/0.5 - 1))
                else :   #special rule for h=infinity / delta `= 0
                    if(Vi > 0.5) :
                        award = stateEV[j]  #Dems sweep this state's EVs for this Vi
                totalEV[k] +=award  #DONE WITH NON-NOISY ELECTION,now do the noisy one
                
                noisyAward = 0.  #default = GOP swept the state in noisy election
                if(beta[k] < 900) :   #we're not doing the h=infinity case. for noisy and non, are we inside step?
                    if noisyVi > 0.50 - delta[k] :  #Dems got at least some nonnoisy EV's
                        noisyAward = stateEV[j]
                        if noisyVi < 0.50 + delta[k] :   #split award for non-noisy election
                            noisyAward = 0.5 * stateEV[j] * (1 + beta[k]*(noisyVi/0.5 - 1))
                else :   #special rule for h=infinity / delta =0
                    if(noisyVi > 0.5) :
                        noisyAward = stateEV[j]  #Dems sweep this state's EVs for this (noisy) Vi
                totalEV_noisy[k] += noisyAward
          
            # we've now awarded the EV's for all states for this simulation and h value
            # CHECK FOR DEM WIN FOR NON-NOISY AND NOISY ELECTIONS
            isDemWin = "no"
            isNoisyDemWin = "no"
            if (totalEV[k] > maxEV/2.) :   #check if Dems won this election for this k=h
                isDemWin = "yes"
                nDemWins[nVus][k] += 1
                nTotDW[k] += 1
            if (totalEV_noisy[k] > maxEV/2.) : #check if Dems won for noisy election
                isNoisyDemWin = "yes"
                nNoisyDemWins[nVus][k] +=1    #now check if election results match ...
                nTotNDW[k] +=1
            if (isDemWin != isNoisyDemWin ) :
                if(isDemWin == "yes") :
                    nMissedDemWins[nVus][k] += 1   #Dems win, but noisy missed
                    nTotMDW[k] += 1
                else :
                    nMissedDemLosses[nVus][k] += 1  #Dems lost, but noisy predicted win
                    nTotMDL[k] += 1

            EVtot[i][k] = totalEV[k]  #store the final total EV for this sim and h value
            sumTotalEV[k] += totalEV[k]    #update the running total for EV's and variance for this Vus

    # COMPLETED ALL SIMS NEAR THIS VALUE OF Vus *** 
    natAvgString = "{:.4f}".format(Vus)
    print(str(noSims)+" simulations done, ready to output stats for this V_US = "+natAvgString)
    # need to calculate EV avg and variance for all h values for this natl Avg
    for k in range(nBetas) :
        localSumVar = 0
        localAvg = sumTotalEV[k] / float(noSims)
        for i in range(noSims) :  #need to loop over sims to compute std dev
            localSumVar += (EVtot[i][k]-localAvg)**2            

    #TIME TO CREATE THE OUTPUT WORKSHEET FOR THIS NOMINAL Vus VALUE
    sheet1 = wb.add_sheet("Vus"+"{:.4f}".format(Vus))
    sheet1.write(0,0,"noSims")  #write the title rows to the output file
    sheet1.write(0,1, "V_US")
    sheet1.write(0,2, "beta=0.5/d")
    sheet1.write(0,3,"true totEVs")
    sheet1.write(0,4,"avg totEVs")
    sheet1.write(0,5,"stdev totEVs")
    sheet1.write(0,6,"DemWins")
    sheet1.write(0,7,"noisy DemW")
    sheet1.write(0,8,"missed DemWs")
    sheet1.write(0,9,"missed DemLs")
    sheet1.write(0,10,"TotDW")
    #sheet1.write(1,10,ALavgDD)   #write out the Alabama stats to remind us which model we used
    sheet1.write(0,11,"tot noisyDW")
    sheet1.write(0,12,"tot missedDW")
    sheet1.write(0,13,"tot missedDL")
    sheet1.write(0,14,stateName[0])
    sheet1.write(1,14,avgDelta[0])  #this is for quality control on which di set we used
    #sheet1.write(2,10,stateName[powerStateNo])
    #sheet1.write(2,11,delta_Vi)
    sheet1.write(nBetas+3,0,"shape k")   #ignore; holdover from gamma distro code
    sheet1.write(nBetas+3,1,kappa)
    sheet1.write(nBetas+4,0,"scale th")  #ignore; holdover from gamma distro code
    sheet1.write(nBetas+4,1,theta)
    sheet1.write(nBetas+5,0,"noise")
    sheet1.write(nBetas+5,1,noise)

    Vus4digit = round(Vus,4)
    for k in range(nBetas) :   # now write the output, one row per simulated h=beta
        avgTotalEV = round(sumTotalEV[k]/noSims,4)
        # avgSD = round((sumVarTotalEV[k] /(noSims-1.0) ) **0.5,4)

        currentRow = k+1
        sheet1.write(currentRow,0, noSims)
        sheet1.write(currentRow,1, Vus4digit)
        sheet1.write(currentRow,2, beta[k])
        #sheet1.write(currentRow,3, nEV_mean[nVus][k])
        sheet1.write(currentRow,4, avgTotalEV)
        sheet1.write(currentRow,5, stdevEV[k])
        sheet1.write(currentRow,6, nDemWins[nVus][k])
        sheet1.write(currentRow,7, nNoisyDemWins[nVus][k])
        sheet1.write(currentRow,8, nMissedDemWins[nVus][k])
        sheet1.write(currentRow,9, nMissedDemLosses[nVus][k])
        sheet1.write(currentRow,10, nTotDW[k])
        sheet1.write(currentRow,11, nTotNDW[k])
        sheet1.write(currentRow,12, nTotMDW[k])
        sheet1.write(currentRow,13, nTotMDL[k])
    outputName = "output"+outName+".xls"
    wb.save(outputName)

# ALL SIMS DONE FOR ALL Vus.  Let's create one more worksheet with some summarized stats
w2 = 0.02      #this is the width of peaked distribution - could be user input
w3 = 0.01      #another peaked distribution
sheet1 = wb.add_sheet("summary")
sheet1.write(nBetas+3,0,"shape k")   #output the parameters so I don't forget 'em
sheet1.write(nBetas+3,1,kappa)
sheet1.write(nBetas+4,0,"scale th")
sheet1.write(nBetas+4,1,theta)
sheet1.write(nBetas+5,0,"noise")
sheet1.write(nBetas+5,1,noise)
sheet1.write(0,0,"nSims")
sheet1.write(1,0,noSims)
sheet1.write(0,1,"n_Vus")
sheet1.write(1,1,numberVus)
sheet1.write(0,3,"h=beta")
# sheet1.write(0,4,"Vustu-mean")
# sheet1.write(0,5,"Vustu-SD")
sheet1.write(0,6,"M1 MErate")  #missed-election rate for model 1 - uniform f(Vus)
sheet1.write(0,7,"M2 MErate")  #and for model 2 peaked f(Vus) and w=0.02
sheet1.write(0,8,"M3 MErate")   #and for model 2, but w=0.01 vs. w=0.02
sheet1.write(0,9,"M2/M1")
sheet1.write(0,10,"M3/M1")
sheet1.write(nBetas+5,7,"M2 w =")
sheet1.write(nBetas+6,7,w2)
sheet1.write(nBetas+5,8,"M3 w =")
sheet1.write(nBetas+6,8,w3)
# must assure normalization for f(Vus) for model 2 for each k value
f2norm = [0.]*nBetas #normalization of pdf for model2.  Should be close to 1
f3norm = [0.]*nBetas
missedRate1 = [0.]*nBetas #will total the missed election rate for model 1
missedRate2 = [0.]*nBetas #same, but for model 2 (peaked distro)
missedRate3 = [0.]*nBetas  #same, but with a different width than model 2
for k in range(nBetas) :
    for nVus in range(numberVus) :
        Vus = minNatlAvg + nVus*deltaVus
        f2norm[k] += deltaVus * norm.pdf(Vus,meanVustu[k],w2)  #running total for model 2's sum(f)
        f3norm[k] += deltaVus * norm.pdf(Vus,meanVustu[k],w3)  #running total for model 2's sum(f)
    print("f2 normalization was "+ str(round(f2norm[k],4) )+" for h = "+str(beta[k]) )
    print("f3 normalization was "+ str(round(f3norm[k],4) )+" for h = "+str(beta[k]) )
    for nVus in range(numberVus) :
        Vus = minNatlAvg + nVus*deltaVus
        missedR = (nMissedDemWins[nVus][k] + nMissedDemLosses[nVus][k] ) / float(noSims)  #local miss rate
        f1Vus = 1.0 / (numberVus - 1.0)  #model 1's weighting is uniform.  Count bookends as half of full wt
        f2Vus = deltaVus * norm.pdf(Vus,meanVustu[k],w2) / f2norm[k]  #weighting factor for model 2 (peaked f(Vus))
        f3Vus = deltaVus * norm.pdf(Vus,meanVustu[k],w3) / f3norm[k] 
        missedRate1[k] += f1Vus * missedR
        missedRate2[k] += f2Vus * missedR
        missedRate3[k] += f3Vus * missedR
for k in range(nBetas) :
    currentRow = k+1
    sheet1.write(currentRow,3,beta[k])
    # sheet1.write(currentRow,4,meanVustu[k])   #dropped from covariance code
    # sheet1.write(currentRow,5,stdevVustu[k])  #dropped from covariance code
    sheet1.write(currentRow,6,missedRate1[k])
    # sheet1.write(currentRow,7,missedRate2[k])  #dropped from covariance code
    # sheet1.write(currentRow,8,missedRate3[k])  #dropped from covariance code
    # sheet1.write(currentRow,9,missedRate2[k]/missedRate1[k])   #dropped
    # sheet1.write(currentRow,10,missedRate3[k]/missedRate1[k])  #dropped
# now write down the state data for posterity
sheet1.write(0,15,"state")
sheet1.write(0,15+1,"st EV")
sheet1.write(0,15+2,"mean di")
sheet1.write(0,15+3,"di stdev")
for j in range(nStates) :
    sheet1.write(j+1,15,stateName[j])
    sheet1.write(j+1,15+1,stateEV[j])
    sheet1.write(j+1,15+2,avgDelta[j])
    sheet1.write(j+1,15+3,stateSD[j])

wb.save(outputName)