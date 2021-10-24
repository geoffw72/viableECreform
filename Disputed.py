
# For EC paper, this code calculates the avg no of disputed states for a given {di}
# the {di} set comes from historical national landscapes (not gamma distro)
# election accuracy at each sampled {di} is based on 1005 noisy elections
# for elections with inaccuracy > 10%, we count the number of responsive states in the distro
# we assume a uniform f(Vus).  We sample across this Vus without binning
# the core election sim code and input code are from WrongWinner.py
# the number of disputed states is a simplified version of hurricane code, where
#       states that are within 1.282 e of being responsive are disputed.  No correlation computation used.
# Note: all my code expresses Vi in terms of Dem's share, while paper uses GOP as candidate 1

from openpyxl import load_workbook  #prepping to read the state data
wb = load_workbook(filename = "state_data.xlsx")
ws = wb.active
import numpy as np
from scipy.stats import norm
import xlwt
from xlwt import Workbook
wb = Workbook()
# minNatlAvg = 0.475
minNatlAvg = input("What is the min. national average fraction vote for Dems, e.g. 0.478? ")
# maxNatlAvg = 0.535 
maxNatlAvg = input("What is the max national average fraction vote for Dems, e.g. 0.538? ")
# numberVus = 61
# numberVus = input("How many Vus to simulate? (e.g. 61) ")
nStates=51  # Excel data in rows B, C, D, E starting in row2
print("number of states = "+str(nStates)+". Time to read in state data.")
nonnoise = [0.]*nStates #initial the election-specific state dev'n from nat'l avg (local to loop)
maxEV = 0   # this will be total US available electoral college votes (global variable)
stateName = ["xx"]*nStates #initializing state's name, no of EVs, mean di, std-dev
stateEV = [0.]*nStates
avgDelta = [0.]*nStates
stateSD = [0.]*nStates
Vi = [0.]*nStates   #we now keep track of each state's Vi for a given sim
inPlay = [0]*nStates  #local to each election -- is this state disputable?  1=yes

for j in range(nStates) :   #Let's read in the data (state name, state EV, avgDelta, stateSD)
    currentRow = str(j+2)  #first row of this file is a header row
    currentStateNameCell = "B"+currentRow
    currentStateName = ws[currentStateNameCell]
    currentStateEVCell = "C" + currentRow
    currentStateEV = ws[currentStateEVCell]
    currentStateADCell = "D" + currentRow
    currentStateAD = ws[currentStateADCell]
    currentStateSDCell = "E" + currentRow
    currentStateSD = ws[currentStateSDCell]
    maxEV += currentStateEV.value  # add this state's EV max to the US max
   # assign this state's properties
    stateName[j] =currentStateName.value
    stateEV[j] = currentStateEV.value
    avgDelta[j] = float(currentStateAD.value)
    stateSD[j] = currentStateSD.value

print("The total number of EV's available in the US is "+str(maxEV))
Sims = input("How many TOTAL simulations to run; e.g. 100000 ")
noSims = int(Sims)
print("Each state's Dem EV's will be a tilted step between 0.5-d and 0.5+d")
nBetas = int(7)
totalEV = [0.]*nBetas #initialize electoral votes for each h

totalRecounts = [0]*nBetas  #running counter of total number of flippable disputes per h value
totalDisputedStates = [0]*nBetas  #running counter of total states embroiled in these disputes
totalFlaggedElections = [0]*nBetas  #number of elections that are possibly disputable
beta=[999,50,20,14,10,6,3]  #beta = h.  h=999 is  approx for current EC.  Pointless to run h=1

noise = input("What is the stdev in noise per state? (0.001, .0017, 0.003, 0.005, 0.01) ")
outName = raw_input("give me a short name for output file, e.g. 2024-0017 ")
p_recount = input("What is the reversal threshold for a dispute? (e.g. 0.10) ")
thresh = norm.ppf(1.-p_recount)  # how far away is disputable (relative to noise level)
strg = str(round(thresh,4))
print("OK, so results within "+strg+" * "+str(noise)+" of responsive will be disputable")
noNoisyElections = input("How many noisy elections to run per close outcome? e.g. 1005 ")
print("I will now run "+str(noSims)+" simulated elections")

for nVus in range(1) : #we are no longer binning results, but I'm too lazy to remove this extra loop
    for i in range(noSims) :  #run a number of simulations near this Vus value
        Vus = minNatlAvg + (maxNatlAvg - minNatlAvg)* float(i)/noSims  #glide thru all Vus's
        if(i/1000 * 1000 == i):  #keep us updated on simulation progress
            print("I am running simulation number "+str(i)+" with Vus = "+str(round(Vus,4)) )
        inPlay=[0.]*nStates   #rezero out this counter for number of disputed states
        # For each sim, assign each state's Vi informed by the f(Vi) distribution
        sumOffset = 0.  #start counter for random drift from sum[EVtot(j)*di(j)] must = 0

        #  **below section uses specified {di,si} set, not a gamma distro
        for j in range(nStates) :  #loop to tentatively assign the delVi's
            nonnoise[j] = np.random.normal(0.0,stateSD[j])  #here, we sample the (di,sigma-i) space
            sumOffset += stateEV[j]*( avgDelta[j]+nonnoise[j] )  #will be used in below normalization
        #Now, tweak the nonnoise to force normalization.  We do not normalize the error set
        for j in range(nStates) :
            nonnoise[j] = nonnoise[j]-sumOffset/maxEV  #correcting for bias in overall state dev'ns
            Vi[j] = Vus + avgDelta[j] + nonnoise[j]
        #  **end of code block to assign the {di} set for this sim

        for k in range(nBetas) :  #loop to compute expected, min and max possible EV's for this h parameter
            maxSwingUp = 0.     #these total how much higher/lower the EV award could be with noise
            maxSwingDown = 0.  #(the potential swings up and down)
            sumVar = 0.    #this will track the variance from responsive states
            nSwingStates = 0.   #these are states inside step, or close enough for the dispute simul'ns
            totalEV[k] = 0       # Reset how many EVs the Dems have before counting Alabama
            for j in range(nStates) : #now loop over states to calc awards for this h value
                inPlay[j] = 0   #default is state would not be relevant for disputed election
                if (beta[k] > 500 ) : #special award algorithm for EC
                    award = 0.
                    if Vi[j] > 0.5 :
                        award = stateEV[j]
                else :  #not an EC case
                    fracAward = 0.5 * (1. + beta[k]*(Vi[j]/0.5 - 1.))
                    award = stateEV[j]* max(0.,min(1.,fracAward))
                # remember that "thresh" was computed from the inputted p_recount
                if (Vi[j] + thresh * noise > 0.50 - 0.5/beta[k] ):  #Dems might get some noisy EV's
                    if (Vi[j] - thresh * noise < 0.50 + 0.5/beta[k]) :   #GOP might get some noisy EV's
                        # likely split award; this state is "in play".  Compute max & min EV awards
                        minVi = Vi[j] - thresh*noise
                        maxVi = Vi[j] + thresh*noise  #approx max and min Vi's
                        minAward = stateEV[j]* max(0.5 * (1. + beta[k]*(minVi/0.5 - 1.)),0. )
                        maxAward = stateEV[j]* min(0.5 * (1. + beta[k]*(maxVi/0.5 - 1.)),1. )
                        # in above two, min/max are to protect for "falling off the step" for this h value
                        maxSwingDown += award - minAward  #update possible swings from this state
                        maxSwingUp += maxAward - award
                        inPlay[j] = 1  #this state would be disputed if this becomes a national dispute
                        sumVar +=  stateEV[j]**2  #measure of variance contribn from this st / (noise * h)
                        nSwingStates += 1.  #update counter on how many potentially in-play states
                    else : # Dems got it all in standard election, and will in noisy election
                            award = stateEV[j]
                totalEV[k] +=award  #update total Dem EVs in this non-noisy election

            # j-loop done, we've now awarded the EV's for all states for this simulation and h value
            noRecountStates = 0  #initialize the number of states involved in a recount
            # to minimize CPU time, only simulate noisy elections if margin is close to flippable
            isDemWin = "no" 
            calcFlip = "no" #flag for running noisy elections
            margin = totalEV[k] - maxEV/2.  #this is positive or negative -- how close we are to a tie for this h
            # I try two tests to look for flippable elections.  If either triggered, run the noisy elections
            #first check follows here - would total possible swing trigger a national flip?
            if(margin > 0.) :  #Dems would win nonnoisy election
                isDemWin = "yes"
                if (margin <= maxSwingDown)  :  # result could swing to GOP (who get 269-269 ties)
                    calcFlip = "yes"
            else :  #(margin <= 0, GOP has at least 269 EVs
                isDemWin = "no"
                if (-1.* margin < maxSwingUp)  :  # result could swing to Dems
                    calcFlip = "yes"
            # begin 2nd check -- is the swing statistically large enough compared to the margin?
            # this check is more relevant with many responsive states
            stDev = noise * beta[k]* sumVar**0.5  #this starts the 2nd check of flippability
            if (stDev < 0.00001) :  #occasionally, no states in zone.  fudge a small stDev to avoid div/zero
                stDev = 0.00001
            ratio = abs(margin / stDev)  # for h<inf, we need to compare the variance to the margin
            chanceFlip = norm.cdf(-1.*ratio)
            if (chanceFlip >= p_recount) :   #is election flippable (finite h test)
                calcFlip = "yes"

            if (calcFlip == "yes") :   #Yes, we need to run the noisy elections to compute flip chance
                totalFlaggedElections[k] += 1  #counter for code health check, not used in output
                strg = str(beta[k])+", "+str(round(margin,3))+", "+str(nSwingStates) 
                print("possible flip.  h, margin, involved states = "+strg)
                noisyWins = 0.  #local counter
                nNE = float(noNoisyElections)
                for nn in range(noNoisyElections) : #run the noisy elections (user had inputted number to run)
                    totalNoisyEV = 0.  #this is our counter for local noisy election total EVs
                    for j in range(nStates) :
                        Vnoisy = Vi[j]+inPlay[j]*np.random.normal(0.0,noise)  #only perturb states that are in play
                        if(beta[k] > 500) : #EC award as special case
                            award = 0.
                            if Vnoisy > 0.5 :
                                award = stateEV[j]
                        else :  #not an EC case
                            fracAward = 0.5 * (1. + beta[k]*(Vnoisy/0.5 - 1.) )
                            award = stateEV[j]* max(0.,min(1.,fracAward) )

                        totalNoisyEV += award 
                    # j loop done on states.  Check national result

                    if(totalNoisyEV > maxEV / 2. ) : #Check if Dems won noisy election
                            noisyWins += 1  #yes, Dem win
                # done running all noisy elections.  Compute fraction won by Dems
                fractionWon = noisyWins / nNE
                # print("fraction of noisy elections won by Dems = "+str(fractionWon))  #debug line; ignore
                if (fractionWon > p_recount) : # the Dems won at least 10pct of the noisy elections
                    if (fractionWon < 1. - p_recount) : #GOP also won at least at 10pct --> election was flippable
                        isRecount = "yes"
                        totalRecounts[k] += 1
                        totalDisputedStates[k] += nSwingStates
            # end of calcFlip loop.  move on to next k = h value

#Time to create the output file
for dummy in range(1) : #only one output sheet; too lazy to eliminate indent
    sheet1 = wb.add_sheet("output")
    sheet1.write(0,0,"noSims")  #write the title rows to the output file
    sheet1.write(0,2, "beta=0.5/d")
    sheet1.write(0,3,"?bl_Elec")  #these are the elections that we TESTED for disputability, but some were <10%
    sheet1.write(0,4,"dispElec")  #number of truly disputable elections (>10% chance of a flip)
    sheet1.write(0,5,"hurrSize")  #this is the average number of states per dispute; the "hurricane" size

    for k in range(nBetas) :   # now write the output, one row per simulated beta
        avgNoDisputedStates= float(totalDisputedStates[k])/float(totalRecounts[k])
        currentRow = k+1
        sheet1.write(currentRow,0, noSims)
        sheet1.write(currentRow,2, beta[k])
        sheet1.write(currentRow,3,totalFlaggedElections[k])
        sheet1.write(currentRow,4, totalRecounts[k])
        sheet1.write(currentRow,5, avgNoDisputedStates)

    outputName = "output"+outName+".xls"
    wb.save(outputName)