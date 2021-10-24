# This code calculates relative state power by perturbing 1 state's result.
# It shares its {di} generation with the "WrongWinner" code

from openpyxl import load_workbook  #prepping to read the state data
wb = load_workbook(filename = "state_data.xlsx")
ws = wb.active
import numpy as np
from scipy.stats import norm
import xlwt
from xlwt import Workbook
wb = Workbook()
nStates=51  # Excel data in rows B, C, D, E starting in row2
outName = raw_input("type in a short filename ")
nS = input("how many sims? e.g. 10000 ")
print("number of states = "+str(nStates)+". Time to read in state data.")
stateName = ["xx"]*nStates #initializing state's name, no of EVs, mean di, std-dev
stateEV = [0.]*nStates
avgDelta = [0.]*nStates
stateSD = [0.]*nStates
maxEV = 0.  #this is global, how many EV's possible in national vote

for j in range(nStates) :   #Let's read in the data (state name, state EV, avgDelta, stateSD)
    currentRow = str(j+2)  #first row of this file is a header row
    currentStateNameCell = "B"+currentRow
    currentStateName = ws[currentStateNameCell]
    currentStateEVCell = "C" + currentRow
    currentStateEV = ws[currentStateEVCell]
    currentStateADCell = "D" + currentRow
    currentStateAD = ws[currentStateADCell]
    currentStateSDCell = "E" + currentRow
    currentStateSD = ws[currentStateSDCell]
    maxEV += currentStateEV.value  # add this state's EV max to the US max
   # assign this state's properties
    stateName[j] =currentStateName.value
    stateEV[j] = currentStateEV.value
    avgDelta[j] = float(currentStateAD.value)
    stateSD[j] = currentStateSD.value

def get_power(powerNo, di,si,EVtot,h,nStates=51,nSims=500000,fVus="uniform",compRange=0.05,w=0.02) :
    #subroutine for state power
    deltaWins = 0.   #this will track outcomes for perturbed Vi's for the state of interest
    minVus = 0.5-compRange/2.  #note that we center the Vus sampling on 0.5 (at least for now)
    maxVus = 0.5+compRange/2.
    dVi = min(0.01,max(0.002,abs(di[powerNo])/10.))  #setting the perturbation to Vi (less for competve states)
    prt = str(powerNo)+","+str(round(di[powerNo],4))+","+str(round(dVi,5)) + "," + str(h)
    print("stateNo, di, dVi, h are ") + prt  #debug print
    sumFactor = 0.  #this will normalize the fVus's
    for i in range(nSims):
        EVus = 0.
        totalEV = 0.
        Vus = minVus + compRange * float(i)/nSims  #we sample all Vus in competitive range.  will weight later
        factor = norm.cdf(abs(Vus-0.5),w)    #NOT SURE IF THIS IS THE RIGHT normalization for peaked distro ***
        if fVus == "uniform" :
            factor = 1.   # or 1./float(nSims)   #flat weighting factor for uniform fVus distro
        sumFactor += factor  #for normalization
        for j in range(nStates) :
            EVus += EVtot[j]  #for summing the total AVAILABLE national EV's
            wobble = np.random.normal(0.0,si[j])  #note, we do not force sum(di*EVtot-i)=0 for this routine
            Vi =di[j]+Vus + wobble
            if(h > 900 ) : #special award rule for EC
                award = 0
                if(Vi > 0.5) :
                    award = EVtot[j]
            else :  #not using EC rules
                    award = EVtot[j]*min( 1,max(0,0.5*(1+h*(Vi/0.5 - 1))) )
            totalEV += award  #update national outcome to include this state's result
            if(j == powerNo) : #are we evaluating the power state?  If yes, perturb and evaluate
                stdAward = award
                Vi_low = Vi - dVi
                Vi_high = Vi + dVi
                if(h > 900 ) : #special award rule for EC
                    highAward = 0.
                    lowAward = 0.
                    if(Vi_low > 0.5) :
                        lowAward = EVtot[j]
                    if(Vi_high > 0.5) :
                        highAward = EVtot[j]
                else :  #not using EC rules
                    highAward = EVtot[j]*min( 1,max(0,0.5*(1+h*(Vi_high/0.5 - 1))) )
                    lowAward = EVtot[j]*min( 1,max(0,0.5*(1+h*(Vi_low/0.5 - 1))) )
        totalEVHigh = totalEV + highAward - stdAward
        totalEVLow = totalEV + lowAward - stdAward  #these two are the national perturbed outcomes
        # prt = str(round(totalEVLow,3))+", "+str(round(totalEVHigh,3))   #print debug
        # print("total EV, high, low are "+str(round(totalEV,4))+", "+prt) #print debug
        if (totalEVHigh > EVus / 2.) : #we have a high-Vi winner
            if (totalEVLow <= EVus / 2.) : #but a lower Vi for power state would trigger a loss
                deltaWins += 0.5*factor  #divide by two for the double disturbance
                # print("we found a perturbed win! "+str(round(totalEVHigh,3))+","+str(round(totalEVLow,3)))
  
    dP_dVi = deltaWins / sumFactor /dVi     # total wins, normalized by weighted number of Sims and perturbn
    return (dP_dVi)

nBetas = int(8)
nDeltas = nBetas
beta=[9999,50,20,14,10,6,3,1]  #beta = h.  h=9999 is flag for current EC
delta = [0]*nDeltas   #delta is convenient shorthand for EV award calcns
IP = [[0.]*nBetas for j in range(nStates) ] # this will be the absolute influence power of each state
RP = [[0.]*nBetas for j in range(nStates) ] # this will be the relative influence power of each state
sumIP = [0.]*nBetas  #counter on total absolute influence power -- will use to normalize IP's to RP's

for k in range(nBetas) :
    delta[k]=0.50/beta[k]
    sumIP[k] = 0.  #zero this counter
    for j in range(nStates) :
        IPi = get_power(j,avgDelta,stateSD,stateEV,beta[k],nStates=nStates,nSims=nS,fVus="uniform",compRange=0.05,w=0.02)
        IP[j][k] = IPi
        sumIP[k] += IPi

# Absolute state powers calculated.  Let's convert to relative power, output results
sheet1 = wb.add_sheet("nS_"+str(nS))
sheet1.write(0,0,"noSims")  #write the title rows to the output file
sheet1.write(1,0,nS)
sheet1.write(0,1,"tot chgd wins")  #title of win differential column.  this unint'ly includes a f(Vustu) factor in it

sheet1.write(0,2, "beta=0.5/d")
for j in range(nStates):
    sheet1.write(0,3+j,stateName[j])
for k in range(nBetas) :
    sheet1.write(k+1,2,beta[k])
    sheet1.write(k+1,1,sumIP[k])   #this is how many total win differentials there were across all sims for all states
    for j in range(nStates) :
        RP[j][k] = IP[j][k] / sumIP[k] # normalizing the influence power
        sheet1.write(k+1,3+j,RP[j][k])

outputName=outName + ".xls"
wb.save(outputName)