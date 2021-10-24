
# For EC paper, this code generates gamma distributions for state di's
#  and then computes the resulting partisan bias vs. linearization parameter
# For this code, we force the states
from openpyxl import load_workbook  #prepping to read the state data
#fileloc = str("C:\Users\wise.gm\source\repos\2-gen-spaced-gammas\2-gen-spaced-gammas")
wb = load_workbook(filename = "state_data.xlsx") #for EV's
ws = wb.active
import numpy as np
import random
from scipy.stats import gamma
import xlwt
from xlwt import Workbook
wb = Workbook()
rd = random.random()
a = input("what is the shape parameter (e.g. 1.0, 1.4, 1.8) ")
b = input("what is the scale parameter (e.g. 0.04, 0.08, 0.12 ")
nSims = input("how many sims? e.g. 10000 ")

outName = str(a)+"_"+str(b)  #for naming the output file
sheet1 = wb.add_sheet("a_b"+outName)  #start the output file
sheet1.write(0,0,"a=k")  #write a and b values to output file image
sheet1.write(1,0,a)
sheet1.write(3,0,"b=q") 
sheet1.write(4,0,b)
sheet1.write(0,2,"EVtot_i")
# sheet1.write(1,1,"d(i)")
nStates = input("How many states? (e.g. 51) ")
maxEV = 0.   #this counter will sum all state EV's (normally 538 total)
stateName = ["xx"]*nStates
stateEV = [0.]*nStates  #initializing state name and EVtot
EV = [0.]*nStates  #initialize awarded EV for each state
order = [0]*nStates #for later shuffling tracking
absDeltaDem = [0.]*nStates   # this will be the absolute values of a gamma-random set of di's
deltaDem = [0.]*nStates #these will be the signed values

def get_Vustu(di,EVtoti,h=999,nSts=51) :  #this is a function to find Vustu for a given {di} set and h
    toler = 0.001
    #now, iterate to find Vustu for this di set.
    toler = 0.001  #how close we have to be to the toss-up national EV total to stop looping
    for i in range(1) :  #used to be across h values, but now we pass this to the function
        stepSize = 0.001 #this is adjustable, our initial step size when we loop-iterate Vus
        dirxn = 1    #this keeps track of which way we should step next
        Vus = 0.5 # initial guess for toss-up point.  Need to reset for each h
        margin = 2*toler #arbitrary start to margin to ensure we enter below while loop
        maxEV = 0.    #total EV's available, usually 538
        sumOffset = 0.  #start counter for random drift from sum[EVtot(j)*di(j)] must = 0
        for j in range(nSts) :  #loop to tentatively assign the delVi's
            sumOffset += EVtoti[j]*( di[j] )  #will be used in below normalization
            maxEV += EVtoti[j]
        #Now, tweak the nonnoise to force normalization.  
        for j in range(nStates) :
            di[j] = di[j]-sumOffset/maxEV  #correcting for bias in overall state dev'ns (if needed)

        while (abs(margin) > toler) :  # loop for: Is Vus close enough to toss-up point?
            guessedVus = Vus  #this will stick if we are at tossup point
            sumVar = 0.
            # nSwingStates = 0
            totEV = 0.  #counter for totalEV's relative to toss-up
            for j in range(nSts) :  #assign Vi's based on normalized noises and guessed Vus
                Vi = guessedVus + di[j]
                award = 0.       # default is GOP won all EVs
                # inPlay[j] = 0   #default is state would not be relevant for disputed election
                if Vi > 0.50 - 0.5/h :  #Dems got at least some nonnoisy EV's
                    fracAward = 0.5 * (1. + h*(Vi/0.5 - 1.))
                    award = EVtoti[j]* max(0.,min(1.,fracAward)) #note, we use h=999 for EC case
                        # inPlay[j] = 1
                        # nSwingStates += 1  #this state is in play; add to running count
                totEV +=award  #update total Dem EVs in this non-noisy election  
        # j-loop done, we've now awarded the EV's for all states for this simulation and h value
            margin = totEV - maxEV/2.  #this is how close we are to a toss-up at the guessed Vus
            if (margin > 0) :   #Vus is too high
                if (dirxn == 1):  #we were going forward.  Reverse and reduce step size; we overshot
                    stepSize = stepSize * 0.5
                    dirxn = -1
                else : #
                    stepSize = stepSize * 1.1  #keep decreasing Vus, take a bigger step
            else :  #our Vus is too low
                if (dirxn == 1) : #Were we going forward? Yes
                    stepSize = stepSize * 1.1  #keep increasing Vus, take a bigger step
                else :  #time to start going forward; we overshot
                    stepSize = stepSize * 0.5
                    dirxn = 1
            Vus = Vus + stepSize * dirxn   #last line in while loop
    return(guessedVus)
    # ** end of defining get_Vustu function

for j in range(nStates) :   #Let's read in the data (state name, state EV), ignoring avgDelta, stateSD
    currentStateName = ws["B"+str(j+2)]
    currentStateEV = ws["C"+str(j+2)]
   # assign this state's properties
    stateName[j] =currentStateName.value
    stateEV[j] = currentStateEV.value
    maxEV += currentStateEV.value  # add this state's EV max to the US max 
    sheet1.write(0,3+j,stateEV[j])
print("I read in the state EV's, they total "+str(maxEV))
print("I will generate a semi-forced gamma distribution for "+str(nStates)+" states")
h = [999,50,20,14,10,6,3]
sheet1.write(0,4+nStates,"h values")
for k in range(6) :
    sheet1.write(1,4+nStates+k,h[k])  #start of 6 columns for later Vustu[k] results

for i in range(nSims) :  #
    sgn = [1.]*nStates #start out with all di's assumed positive, will flip some later
    for j in range(nStates) :   #let's go back to normal order before we shuffle
        order[j] = j  #we will shuffle the order later
    LG_EV = 0.
    for j in range(nStates) :
        qmin = (j+0.01)/float(nStates)  #not 0.00 to curtail the left of the distro
        qmax = (j+0.99)/float(nStates)  #not 1.00 to curtail the right tail of the distro
        aDDmin = b*gamma.ppf(qmin,a)
        aDDmax = b*gamma.ppf(qmax,a)
        qq = (j+0.5)/float(nStates) # center of relative location of this state on the polazn cont'm
        # absDeltaDem[j] = aDDmin + (aDDmax - aDDmin)*random.random()  # d(i) will be somewhere in the gamme interval
        absDeltaDem[j] = b*gamma.ppf(qq,a)+np.random.normal(0.0,0.2*(aDDmax-aDDmin))  #0.2 is arbitrary param
        # here, we make it more likely that the assigned di's are in the center of their "apartments" but allow range
        LG_EV += absDeltaDem[j]*stateEV[j]  #counter for sum(di*EVtot_i) --
    random.shuffle(order)  #this is the order of the states

    jj = 0  #counter in below loop.  wouldn't kick out if I set up as a "for" loop
    while LG_EV > 0. : #in this loop, we will flip di signs until we're past even
        oldLG_EV = LG_EV  #this will determine if we reject the last "flip"
        sgn[order[jj]]=-1.  #this state is now of opposite charge
        LG_EV += -2.*absDeltaDem[order[jj]]*stateEV[order[jj]]
        jj += 1
        #end of while loop
    # if (oldLG_EV < -1.* LG_EV) :  # check if last flip put us further away from zero
    #    sgn[order[jj-1]] = 1.   #yes, too far.  revert last flip
    flip = np.sign(np.random.normal(0,1))  #arbitrarily decide whether re-signed states are Dem or GOP
    for jj in range(nStates) :
        j = order.index(jj)   # to invert the order shuffling
        deltaDem[j]=flip*sgn[j]*absDeltaDem[j]
        EV[j] = stateEV[j]*deltaDem[j]  #this is actually LG EV award - that at Vi=Vus
    #now, let's re-center the di's
    adjust = np.sum(EV)/maxEV
    for j in range(nStates) :
        deltaDem[j] -= adjust
        EV[j] = stateEV[j]*deltaDem[j]  #this adjustment forces sum(di*EVi) to zero at Vus=0.5

    sheet1.write(3+i,1,i+1)
    sheet1.write(3+i,2,"d(i)-"+str(i+1))    #title this column of di's in output file
    for j in range(nStates) :
        sheet1.write(3+i,3+j,deltaDem[j])
    # sheet1.write(nStates+3,3+i, "EVtot")
    #sheet1.write(nStates+4,3+i,np.sum(EV))
    for k in range(6) :
        Vustu = get_Vustu(deltaDem,stateEV,h[k],nStates)
        # print("the toss-up Vus for this {di} is "+str(round(Vustu,5))+" at h = "+str(h[k]))
        sheet1.write(3+i,4+nStates+k,Vustu)
# sheet1.write(nStates+3,1,np.average(EV))
# sheet1.write(nStates+4,1,np.std(EV))
wb.save(outName+".xls")